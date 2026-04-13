import os
import json
import urllib.parse
import re
import html as html_module
import re
import html as html_module
from pathlib import Path
from rest_framework.response import Response
from rest_framework import viewsets
from django.http import JsonResponse
from django.apps import apps
import csv
from django.http import StreamingHttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils.html import strip_tags
from html.parser import HTMLParser

from django.shortcuts import render, get_object_or_404
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt

from rest_framework import viewsets
from rest_framework_gis.filters import InBBoxFilter
from rest_framework.filters import SearchFilter
from django.views.generic.detail import DetailView
from rest_framework.views import APIView
from rest_framework import status
from .forms import AnonymousMemorialForm
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.contrib.contenttypes.models import ContentType
from django.contrib.gis.db.models.functions import Distance
from django.contrib.gis.measure import D
from django.utils.translation import gettext as _

from .models import (
    PartisanMemorial,
    CroatianPartisanMemorial,
    PartisanHospital,
    PartisanNaming,
    PartisanPointsWithoutMemorial,
    OtherMemorials,
    ConnectedExternalEntry,
    MemorialImage,
)

def html_to_text_preserving_breaks(content: str) -> str:
    """Convert HTML to plain text while preserving logical line breaks.

    - <br> -> newline
    - block closers (</p>, </div>, </li>, </h1>..</h6>, </tr>) -> newline
    - remove remaining tags
    - unescape HTML entities
    - normalize CRLF/CR to \n and trim trailing spaces around lines
    """
    if not content:
        return ""
    text = str(content)
    # Normalize common break tags to newlines
    text = re.sub(r"(?i)<br\s*/?>", "\n", text)
    # Put a newline at the end of common block-level elements
    text = re.sub(r"(?i)</(p|div|li|h[1-6]|tr)>", "\n", text)
    # Remove opening tags of those blocks (we already added breaks on closing)
    text = re.sub(r"(?i)<(p|div|li|h[1-6]|tr)(\s[^>]*)?>", "", text)
    # Remove all remaining tags
    text = strip_tags(text)
    # Unescape HTML entities (e.g., &nbsp;)
    text = html_module.unescape(text)
    # Normalize newlines
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    # Trim extra spaces around lines but keep explicit line breaks
    lines = [ln.strip() for ln in text.split("\n")]
    # Drop accidental consecutive empty lines to at most one
    normalized_lines = []
    prev_empty = False
    for ln in lines:
        is_empty = (ln == "")
        if is_empty and prev_empty:
            continue
        normalized_lines.append(ln)
        prev_empty = is_empty
    return "\n".join(normalized_lines)


class _LinkReplacingParser(HTMLParser):
    """Turn HTML into text while rendering anchors as 'text (href)'.

    Also preserves logical line breaks for <br> and common block-level tags.
    """
    _BLOCK_TAGS = {f'h{i}' for i in range(1, 7)} | {'p', 'div', 'li', 'tr'}

    def __init__(self):
        super().__init__()
        self._out = []
        self._in_a = False
        self._current_href = None

    def handle_starttag(self, tag, attrs):
        t = tag.lower()
        if t == 'br':
            self._out.append('\n')
        if t == 'a':
            self._in_a = True
            self._current_href = None
            for k, v in attrs:
                if k.lower() == 'href':
                    self._current_href = v
                    break

    def handle_endtag(self, tag):
        t = tag.lower()
        if t in self._BLOCK_TAGS:
            self._out.append('\n')
        if t == 'a':
            self._in_a = False
            self._current_href = None

    def handle_data(self, data):
        if not data:
            return
        if self._in_a and self._current_href:
            self._out.append(f"{data} ({self._current_href})")
        else:
            self._out.append(data)

    def get_text(self) -> str:
        text = ''.join(self._out)
        text = html_module.unescape(text)
        text = text.replace('\r\n', '\n').replace('\r', '\n')
        lines = [ln.strip() for ln in text.split('\n')]
        normalized_lines = []
        prev_empty = False
        for ln in lines:
            is_empty = (ln == '')
            if is_empty and prev_empty:
                continue
            normalized_lines.append(ln)
            prev_empty = is_empty
        return '\n'.join(normalized_lines)


def html_to_text_with_links(html: str) -> str:
    """Convert HTML to plain text and render links as 'text (url)'."""
    try:
        parser = _LinkReplacingParser()
        parser.feed(str(html))
        parser.close()
        return parser.get_text()
    except Exception:
        return html_to_text_preserving_breaks(html)


DETAIL_MODELS = (
    PartisanMemorial,
    CroatianPartisanMemorial,
    PartisanHospital,
    PartisanNaming,
    PartisanPointsWithoutMemorial,
    OtherMemorials,
)

TRANSLATABLE_MODEL_SLUGS = (
    'partisanmemorial',
    'croatianpartisanmemorial',
    'partisanhospital',
    'partisannaming',
    'partisanpointswithoutmemorial',
    'othermemorials',
)


class MemorialPublicDetailView(View):
    """Public detail page for a single memorial-like point with nearby items."""

    NEARBY_RADIUS_KM = 10
    NEARBY_LIMIT_PER_MODEL = 12
    NEARBY_LIMIT_TOTAL = 20

    @staticmethod
    def _normalize_soft_wrapped_text(text: str) -> str:
        """Collapse wrapped lines and suppress repeated blank lines."""
        if not text:
            return text
        text = text.replace('\r\n', '\n').replace('\r', '\n')
        text = re.sub(r'(?<!\n)\n(?!\n)', ' ', text)
        text = re.sub(r'[ \t]{2,}', ' ', text)
        text = re.sub(r'\n{2,}', '\n', text)
        return text.strip()

    def _slug_to_model_name(self, model_slug: str) -> str:
        slug = (model_slug or '').lower()
        for canonical in TRANSLATABLE_MODEL_SLUGS:
            translated = str(_(canonical)).lower()
            if slug in {canonical, translated}:
                return canonical
        return slug

    def _model_name_to_slug(self, model_name: str) -> str:
        return str(_((model_name or '').lower()))

    def _resolve_model(self, model_slug):
        model_name = self._slug_to_model_name(model_slug)
        try:
            model = apps.get_model('zemljevid', model_name)
        except LookupError:
            return None
        if model not in DETAIL_MODELS:
            return None
        return model

    def _value_to_text(self, value):
        import datetime

        if value is None:
            return ''
        if isinstance(value, (list, tuple, set)):
            cleaned = [str(v).strip() for v in value if v not in (None, '', []) and str(v).strip() != '']
            return ', '.join(cleaned)
        if isinstance(value, (datetime.date, datetime.datetime, datetime.time)):
            return value.isoformat()
        if hasattr(value, 'all'):
            return ', '.join(str(v) for v in value.all())
        if isinstance(value, str):
            if ('<' in value and '>' in value) and (
                ('<br' in value.lower())
                or ('<p' in value.lower())
                or ('<div' in value.lower())
                or ('<a ' in value.lower())
            ):
                return self._normalize_soft_wrapped_text(html_to_text_with_links(value))
            return self._normalize_soft_wrapped_text(value)
        return str(value)

    def _build_external_url(self, pattern, ext_id):
        if not pattern:
            return None
        ext_id = ext_id or ''
        if '[ID]' in pattern:
            return pattern.replace('[ID]', ext_id)
        if ext_id:
            if pattern.endswith('/'):
                return pattern + ext_id
            return pattern.rstrip('/') + '/' + ext_id
        return pattern

    def get(self, request, *args, **kwargs):
        model_slug = kwargs.get('model_slug')
        object_id = kwargs.get('object_id')

        if not model_slug or not object_id:
            return JsonResponse({"error": "Missing model_name or object_id"}, status=400)

        model = self._resolve_model(model_slug)
        if model is None:
            return JsonResponse({"error": "Model not found"}, status=404)

        obj = get_object_or_404(model, pk=object_id)

        display_fields = []
        for field in model._meta.fields:
            if field.name in {'id', 'geom', 'hidden', 'remarks'}:
                continue
            raw_value = getattr(obj, field.name, None)

            if getattr(field, 'choices', None):
                raw_value = dict(field.flatchoices).get(raw_value, raw_value)

            value = self._value_to_text(raw_value)
            if value in ('', None):
                continue
            display_fields.append({
                'key': str(field.verbose_name).capitalize(),
                'value': value,
            })

        for field in model._meta.many_to_many:
            raw_value = getattr(obj, field.name, None)
            value = self._value_to_text(raw_value)
            if value in ('', None):
                continue
            display_fields.append({
                'key': str(field.verbose_name).capitalize(),
                'value': value,
            })

        nearby = []
        if obj.geom:
            for nearby_model in DETAIL_MODELS:
                qs = nearby_model.objects.filter(geom__isnull=False)
                if hasattr(nearby_model, 'hidden'):
                    qs = qs.filter(hidden=False)
                if nearby_model == model:
                    qs = qs.exclude(pk=obj.pk)

                qs = (
                    qs.annotate(distance=Distance('geom', obj.geom))
                    .filter(geom__distance_lte=(obj.geom, D(km=self.NEARBY_RADIUS_KM)))
                    .order_by('distance')[: self.NEARBY_LIMIT_PER_MODEL]
                )

                for item in qs:
                    distance_m = getattr(item, 'distance', None)
                    item_model_name = item._meta.model_name
                    nearby.append({
                        'model_name': item_model_name,
                        'model_slug': self._model_name_to_slug(item_model_name),
                        'model_verbose': str(item._meta.verbose_name),
                        'id': item.pk,
                        'name': item.name or f"#{item.pk}",
                        'distance_km': round((distance_m.m if distance_m else 0) / 1000, 2),
                        'lat': item.geom.y if item.geom else None,
                        'lng': item.geom.x if item.geom else None,
                    })

            nearby = sorted(nearby, key=lambda x: x['distance_km'])[: self.NEARBY_LIMIT_TOTAL]

        content_type = ContentType.objects.get_for_model(model)
        images = list(
            MemorialImage.objects.filter(content_type=content_type, object_id=obj.pk)
            .order_by('order', 'id')[:12]
        )

        connected_entries = []
        ext_entries = ConnectedExternalEntry.objects.select_related('external_project').filter(
            content_type=content_type,
            object_id=obj.pk,
        )
        for entry in ext_entries:
            connected_entries.append({
                'project_name': entry.external_project.name,
                'external_id': entry.external_id,
                'url': self._build_external_url(
                    getattr(entry.external_project, 'url', None),
                    entry.external_id,
                ),
            })

        context = {
            'model_name': model._meta.model_name,
            'model_slug': self._model_name_to_slug(model._meta.model_name),
            'model_verbose_name': str(model._meta.verbose_name),
            'object_id': obj.pk,
            'object_name': obj.name or f"#{obj.pk}",
            'display_fields': display_fields,
            'lat': obj.geom.y if obj.geom else None,
            'lng': obj.geom.x if obj.geom else None,
            'nearby': nearby,
            'images': images,
            'connected_entries': connected_entries,
            'nearby_radius_km': self.NEARBY_RADIUS_KM,
            'mini_map_data_json': {
                'lat': obj.geom.y if obj.geom else None,
                'lng': obj.geom.x if obj.geom else None,
                'name': obj.name or f"#{obj.pk}",
                'radius_m': self.NEARBY_RADIUS_KM * 1000,
            },
            'nearby_points_json': [
                {
                    'lat': item['lat'],
                    'lng': item['lng'],
                    'name': item['name'],
                    'distance_km': item['distance_km'],
                }
                for item in nearby
                if item.get('lat') is not None and item.get('lng') is not None
            ],
        }
        return render(request, 'generic_detail.html', context)

# class GalleryView(View):
#     def get(self, request, *args, **kwargs):
#         table_id = request.GET.get('table_id')
#         fid = request.GET.get('fid')

#         if not table_id or not fid:
#             return JsonResponse({"error": "Missing table_id or fid"}, status=400)

#         # Use the ImageListViewSet to get the list of images
#         image_list_view_set = ImageListViewSet()
#         response = image_list_view_set.list(request)
#         images = response.data

#         return render(request, 'gallery.html', {'images': images})
    
class GenericDetailView(View):
    def get(self, request, *args, **kwargs):
        model_name = kwargs.get('model_name')
        object_id = kwargs.get('object_id')

        if not model_name or not object_id:
            return JsonResponse({"error": "Missing model_name or object_id"}, status=400)

        try:
            model = apps.get_model('zemljevid', model_name)
        except LookupError:
            return JsonResponse({"error": "Model not found"}, status=404)

        obj = get_object_or_404(model, pk=object_id)
        fields = {field.name: getattr(obj, field.name) for field in model._meta.get_fields()}

        return render(request, 'generic_detail.html', {'model_name': model_name, 'fields': fields})

class GeopediaDetailView(DetailView):
    model = None  # This will be set dynamically
    template_name = 'generic_detail.html'
    context_object_name = 'entry'
    extra_context = {}

    def get_object(self, queryset=None):
        # Get the model name and object ID from the URL
        model_name = self.kwargs.get('model_name')
        object_id = self.kwargs.get('object_id')

        # Get the model class dynamically
        try:
            self.model = apps.get_model('zemljevid', model_name)
        except LookupError:
            return JsonResponse({"error": "Model not found"}, status=404)

        # Fetch the object using the object ID
        obj = get_object_or_404(self.model, pk=object_id)
        return obj
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['model_name'] = self.kwargs.get('model_name')
        context['object_id'] = self.kwargs.get('object_id')
        context['fields'] = {field.name: getattr(self.object, field.name) for field in self.model._meta.get_fields()}
        return context

class ExportTableCSVView(View):
    def get(self, request, *args, **kwargs):
        model_name = kwargs.get('model_name')
        if not model_name:
            return JsonResponse({"error": "Missing model_name"}, status=400)
        try:
            model = apps.get_model('zemljevid', model_name)
        except LookupError:
            return JsonResponse({"error": "Model not found"}, status=404)
        # Query all objects
        queryset = model.objects.all()
        # Get field names
        field_names = [field.name for field in model._meta.fields]
        # Streaming response for large tables
        def row_generator():
            yield ','.join(field_names) + '\n'
            for obj in queryset.iterator():
                row = [str(getattr(obj, field, '')) for field in field_names]
                yield ','.join(row) + '\n'
        response = StreamingHttpResponse(row_generator(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{model_name}.csv"'
        return response

class ExportTableXLSXView(View):
    def get(self, request, *args, **kwargs):
        model_name = kwargs.get('model_name')
        if not model_name:
            return JsonResponse({"error": "Missing model_name"}, status=400)
        try:
            model = apps.get_model('zemljevid', model_name)
        except LookupError:
            return JsonResponse({"error": "Model not found"}, status=404)
        queryset = model.objects.all()
        # Skip geom field
        field_names = [field.name for field in model._meta.fields if field.name != 'geom']
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = model_name

        ws.append(field_names)
        # Import once (avoid doing it in the inner loop)
        try:
            from tinymce.models import HTMLField
            has_tinymce = True
        except Exception:
            HTMLField = None  # type: ignore
            has_tinymce = False

        for obj in queryset.iterator():
            row_values = []
            for field in field_names:
                value = getattr(obj, field, '')
                # If field is datetime, show only date in YYYY-MM-DD format
                import datetime
                if isinstance(value, datetime.datetime):
                    value = value.date().isoformat()
                elif isinstance(value, datetime.date):
                    value = value.isoformat()
                # Remove timezone info for datetime/time objects
                if hasattr(value, 'tzinfo') and value.tzinfo is not None:
                    value = value.replace(tzinfo=None)
                # Handle HTML fields (TinyMCE) specially: always render links as 'text (url)'
                if has_tinymce:
                    try:
                        model_field = model._meta.get_field(field)
                    except Exception:
                        model_field = None
                    if model_field is not None and HTMLField and isinstance(model_field, HTMLField):
                        raw_html = value or ""
                        value = html_to_text_with_links(raw_html)
                # Fallback: if value is a string that looks like HTML, preserve breaks and render links as 'text (url)'
                if isinstance(value, str) and ('<' in value and '>' in value) and (('<br' in value.lower()) or ('<p' in value.lower()) or ('<div' in value.lower()) or ('<a ' in value.lower())):
                    value = html_to_text_with_links(value)

                row_values.append(value)

            # Append the values
            ws.append(row_values)

        # Set overall font to Arial 12 and wrap cells that contain newlines
        from openpyxl.styles import Font, Alignment
        arial_font = Font(name='Arial', size=12)        
        for row in ws.iter_rows():
            for cell in row:
                cell.font = arial_font
                if isinstance(cell.value, str) and ('\n' in cell.value):
                    cell.alignment = Alignment(wrap_text=True)
                
        # Auto-size columns
        for col_num, column_title in enumerate(field_names, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = max(10, len(column_title) + 2)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{model_name}.xlsx"'
        wb.save(response)
        return response

class ExportTableView(View):
    def get(self, request, *args, **kwargs):
        model_name = kwargs.get('model_name')
        if not model_name:
            return JsonResponse({"error": "Missing model_name"}, status=400)
        export_url = request.build_absolute_uri(f'/export_csv/{model_name}/')
        return render(request, 'export_table.html', {'export_url': export_url, 'model_name': model_name})

def missing_memorial_view(request):
    if request.method == 'POST':
        form = AnonymousMemorialForm(request.POST)
        if form.is_valid():
            form.save()
            return render(request, 'zemljevid/missing_memorial.html', {'form': None, 'success': True})
    else:
        form = AnonymousMemorialForm()
    return render(request, 'zemljevid/missing_memorial.html', {'form': form})